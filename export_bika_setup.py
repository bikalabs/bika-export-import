# usage: interpreter [-h] [-s SITEPATH] [-u USERNAME] [-o OUTPUTFILE]
#
# Export bika_setup into an Open XML (XLSX) workbook
#
# optional arguments:
#   -h, --help     show this help message and exit
#   -s SITEPATH    full path to site root (default: Plone)
#   -u USERNAME    zope admin username (default: admin)
#   -o OUTPUTFILE  output zip file name (default: SITEPATH.zip)
#
# This script is meant to be run with zopepy or bin/instance. See
# http://docs.plone.org/develop/plone/misc/commandline.html for details.

import argparse
import os
import sys
import tempfile
import zipfile
import shutil

from AccessControl.SecurityManagement import newSecurityManager
from bika.lims.catalog import getCatalog
from Products.Archetypes import Field
from Products.CMFCore.interfaces import ITypeInformation
from Products.CMFCore.utils import getToolByName
import openpyxl

# def excepthook(typ, value, tb):
#     import pdb, traceback
#     traceback.print_exception(typ, value, tb)
#     pdb.pm()
# sys.excepthook = excepthook

export_types = [
    'Client',
    'Contact',
    'ARPriority',
    'AnalysisProfile',
    'ARTemplate',
    'AnalysisCategory',
    'AnalysisService',
    'AnalysisSpec',
    'AttachmentType',
    'BatchLabel',
    'Calculation',
    'Container',
    'ContainerType',
    'Department',
    'Instrument',
    'InstrumentCalibration',
    'InstrumentCertification',
    'InstrumentMaintenanceTask',
    'InstrumentScheduledTask',
    'InstrumentType',
    'InstrumentValidation',
    'LabContact',
    'LabProduct',
    'Manufacturer',
    'Method',
    'Preservation',
    'ReferenceDefinition',
    'SampleCondition',
    'SampleMatrix',
    'StorageLocation',
    'SamplePoint',
    'SampleType',
    'SamplingDeviation',
    'SRTemplate',
    'SubGroup',
    'Supplier',
    'SupplierContact',
    'WorksheetTemplate',
]

# fieldnames that are never exported
ignore_fields = [
    # dublin
    'constrainTypesMode',
    'locallyAllowedTypes',
    'immediatelyAddableTypes',
    'subject',
    'relatedItems',
    'location',
    'language',
    'effectiveDate',
    'modification_date',
    'expirationDate',
    'creators',
    'contributors',
    'rights',
    'allowDiscussion',
    'excludeFromNav',
    'nextPreviousEnabled',
]


class Main:
    def __init__(self, args):
        self.args = args
        # pose as user
        self.user = app.acl_users.getUserById(args.username)
        newSecurityManager(None, self.user)
        # get portal object
        self.portal = app.unrestrictedTraverse(args.sitepath)

        self.proxy_cache = {}

    def __call__(self):
        """Export entire bika site
        """
        self.tempdir = tempfile.mkdtemp()
        # Export into tempdir
        self.wb = openpyxl.Workbook()
        self.export_laboratory()
        self.export_bika_setup()
        for portal_type in export_types:
            self.export_portal_type(portal_type)
        self.wb.save(os.path.join(self.tempdir, 'setupdata.xlsx'))
        # Create zip file
        zf = zipfile.ZipFile(self.args.outputfile, 'w', zipfile.ZIP_DEFLATED)
        for fname in os.listdir(self.tempdir):
            zf.write(os.path.join(self.tempdir, fname), fname)
        zf.close()
        # Remove tempdir
        shutil.rmtree(self.tempdir)

    def get_catalog(self, portal_type):
        # grab the first catalog we are indexed in
        at = getToolByName(self.portal, 'archetype_tool')
        return at.getCatalogsByType(portal_type)[0]

    def get_fields(self, schema):
        fields = []
        for field in schema.fields():
            if field.getName() in ignore_fields:
                continue
            if Field.IComputedField.providedBy(field):
                continue
            fields.append(field)
        return fields

    def write_dict_field_values(self, instance, field):
        value = field.get(instance)
        if type(value) == dict:
            value = [value]
        keys = value[0].keys()
        # Create or obtain sheet for this field type's values
        sheetname = '%s_values' % field.type
        sheetname = sheetname[:31]
        if sheetname in self.wb:
            ws = self.wb[sheetname]
        else:
            ws = self.wb.create_sheet(title=sheetname)
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1
            ws.cell(column=1, row=1).value = "id"
            ws.cell(column=2, row=1).value = "field"
            for col, key in enumerate(keys):
                cell = ws.cell(column=col + 3, row=1)
                cell.value = key
        nr_rows = len(ws.rows) + 1
        for row, v in enumerate(value):
            if not any(v.values()):
                break
            # source id/field
            ws.cell(column=1, row=nr_rows + row).value = instance.id
            ws.cell(column=2, row=nr_rows + row).value = field.getName()
            for col, key in enumerate(keys):
                c_value = v.get(key, '')
                ws.cell(column=col + 3, row=nr_rows + row).value = c_value

        return sheetname

    def write_reference_values(self, instance, field):
        values = field.get(instance)
        # Create or obtain sheet for this relationship
        sheetname = field.relationship[:31]
        if sheetname in self.wb:
            ws = self.wb[sheetname]
        else:
            ws = self.wb.create_sheet(title=sheetname)
            ws.cell(column=1, row=1).value = "Source"
            ws.cell(column=2, row=1).value = "Target"
        nr_rows = len(ws.rows) + 1
        for row, value in enumerate(values):
            ws.cell(column=1, row=nr_rows + row).value = instance.id
            ws.cell(column=2, row=nr_rows + row).value = value.id
        return sheetname

    def get_extension(self, mimetype):
        """Return first extension for mimetype, if any is found.
        If no extension found, return ''
        """
        mr = getToolByName(self.portal, "mimetypes_registry")
        extension = ''
        for ext, mt in mr.extensions.items():
            if mimetype == mt:
                extension = ext
        return extension

    def mutate(self, instance, field):
        value = field.get(instance)
        # Booleans are special; we'll str and return them.
        if value is True or value is False:
            return str(value)
        # Zero is special: it's false-ish, but the value is important.
        if value is 0:
            return 0
        # Other falsish values make empty cells.
        if not value:
            return ''
        # Date fields get stringed to rfc8222
        if Field.IDateTimeField.providedBy(field):
            return value.rfc822() if value else None
        # TextField implements IFileField, so we must handle it
        # before IFileField. It's just returned verbatim.
        elif Field.ITextField.providedBy(field):
            return value
        # Files get saved into tempdir, and the cell content is the filename
        elif Field.IFileField.providedBy(field):
            if not value.size:
                return ''
            extension = self.get_extension(value.content_type)
            filename = value.filename if value.filename \
                else instance.id + '-' + field.getName() + "." + extension
            of = open(os.path.join(self.tempdir, filename), 'wb')
            of.write(value.data)
            of.close()
            return filename
        elif Field.IReferenceField.providedBy(field):
            if field.multiValued:
                return self.write_reference_values(instance, field)
            else:
                return value.Title()
        elif Field.ILinesField.providedBy(field):
            return "\n".join(value)
        # depend on value of field, to decide mutation.
        else:
            value = field.get(instance)
            # Dictionaries or lists of dictionaries
            if type(value) == dict \
                    or (type(value) in (list, tuple)
                        and type(value[0]) == dict):
                return self.write_dict_field_values(instance, field)
            else:
                return value

    def export_laboratory(self):
        instance = self.portal.bika_setup.laboratory
        ws = self.wb.create_sheet(title='Laboratory')
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        fields = self.get_fields(instance.schema)
        for row, field in enumerate(fields):
            ws.cell(column=1, row=row + 1).value = field.getName()
            value = self.mutate(instance, field)
            ws.cell(column=2, row=row + 1).value = value

    def export_bika_setup(self):
        instance = self.portal.bika_setup
        ws = self.wb.create_sheet(title='BikaSetup')
        fields = self.get_fields(instance.schema)
        for row, field in enumerate(fields):
            ws.cell(column=1, row=row + 1).value = field.getName()
            value = self.mutate(instance, field)
            ws.cell(column=2, row=row + 1).value = value

    def export_portal_type(self, portal_type):
        catalog = self.get_catalog(portal_type)
        brains = catalog(portal_type=portal_type)
        if not brains:
            print "No objects of type %s found in %s" % (portal_type, catalog)
            return
        ws = self.wb.create_sheet(title=portal_type)
        # Write headers
        instance = brains[0].getObject()
        fields = self.get_fields(instance.schema)
        headers = ['path', 'uid']
        headers += [f.getName() for f in fields]
        for col, header in enumerate(headers):
            ws.cell(column=col + 1, row=1).value = header
        # Write values
        portal_path = '/'.join(self.portal.getPhysicalPath())
        for row, brain in enumerate(brains):
            instance = brain.getObject()
            # path
            path = '/'.join(instance.getPhysicalPath()[:-1])
            ws.cell(column=1, row=row + 2).value = path.replace(portal_path, '')
            # uid
            ws.cell(column=2, row=row + 2).value = instance.UID()
            # then schema field values
            for col, field in enumerate(fields):
                value = self.mutate(instance, field)
                ws.cell(column=col + 3, row=row + 2).value = value


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Export bika_setup into an Open XML (XLSX) workbook',
        epilog='This script is meant to be run with zopepy or bin/instance. See'
               ' http://docs.plone.org/develop/plone/misc/commandline.html for'
               ' details.'
    )
    parser.add_argument(
        '-s',
        dest='sitepath',
        default='Plone',
        help='full path to site root (default: Plone)')
    parser.add_argument(
        '-u',
        dest='username',
        default='admin',
        help='zope admin username (default: admin)')
    parser.add_argument(
        '-o',
        dest='outputfile',
        default='',
        help='output zip file name (default: SITEPATH.zip)')
    args, unknown = parser.parse_known_args()
    if args.outputfile == '':
        args.outputfile = args.sitepath + ".zip"

    main = Main(args)
    main()
